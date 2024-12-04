import csv, time, os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

# path settings
current_year = date.today().year
last_year = current_year - 1
homedir = os.path.expanduser('~')
path_script = os.path.join(os.path.dirname(__file__))
path_year = os.path.join(os.path.dirname(__file__), str(last_year) + "_oapen")
path_input_output = os.path.join(path_script, 'input-output')

print("This script will check if any new countries were added to the Oapen, Jstor, and Project Muse readership data.")
print()
print("To continue, you need the rd_jstor.xlsx and rd_muse.xlsx files, as well as the countries_combined.csv file.")
print("Place these in the following directory:")
print(path_input_output)
print()
input("Press Enter to continue. \n")

# check if there is a combined countries file present
countries_combined_file = os.path.join(path_input_output, "countries_combined.csv")
try:
    with open(countries_combined_file, 'r', newline='', encoding='utf-16') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_NONNUMERIC)
        existing_merged_file = csv.DictReader(file)
        countries_oapen_reference = []
        countries_jstor_reference = []
        countries_muse_reference = []
        # write the contents for each column to lists per database
        for row in existing_merged_file:
            countries_oapen_reference.append(row['countries_oapen'])
            countries_jstor_reference.append(row['countries_jstor'])
            countries_muse_reference.append(row['countries_muse'])
        file.close()
    print()
    print("countries_combined.csv file found. Parsing file.")
except FileNotFoundError:
    print()
    print("No existing countries_combined.csv file found in the directory.")
    print("Please place the countries_combined.csv file in the right location and restart the script.")

# remove empty strings from lists
print()
print("Removing empty strings from lists.")
while '' in countries_oapen_reference:
    countries_oapen_reference.remove('')
while '' in countries_jstor_reference:
    countries_jstor_reference.remove('')
while '' in countries_muse_reference:
    countries_muse_reference.remove('')

# selenium driver settings
print("Starting Selenium webdriver to download the oapen country list.")
chrome_options = Options()
chrome_options.add_argument("--disable-search-engine-choice-screen")
prefs = {"download.default_directory":path_input_output}
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

# specify the url to use (oapen)
url = 'https://dashboard.oapen.org/#/dashboard'

# initialise selenium driver
print("Opening oapen website.")
driver.maximize_window()
driver.get(url)

# login to oapen
# if the username and/org password ever change, you can edit them here
print("Logging in to oapen.")
username = driver.find_element(By.XPATH, '//*[@id="username"]')
username.clear()
# username redacted
username.send_keys("")
password = driver.find_element(By.XPATH, '//*[@id="password"]')
password.clear()
# password redacted
password.send_keys("")
login_button = driver.find_element(By.XPATH, '/html/body/div/table/tbody/tr/td/form/div/table/tbody/tr[3]/td/button')
login_button.click()

# select per month and country view
print("Selecting 'per month and country' view.")
wait = WebDriverWait(driver, 10)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/main/div/div/div/header/div/button/span/i'))).click()
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div[2]/main/div/div/div/aside/div[1]/div/div[3]/div[2]/div'))).click()

# download the .csv file
print("Downloading .csv file.")
download_button = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[contains(@id,"json-to-csv-")]/button')))
download_button.click()

time.sleep(2)

# close the selenium webdriver
print("Closing Selenium Webdriver")
driver.quit()

oapen_download_file = os.path.join(path_input_output, "monthly_requests_per_country.csv")
if os.path.isfile(oapen_download_file) == True:
    print()
    print("Oapen download file found.")
else:
    print()
    print("WARNING: No Oapen download file found. The script will continue to run with only the Jstor and Project Muse data.")
    print("If you want to run it properly, however, try again.")

# open the oapen file in universal line ending mode 
print()
print("Making a list of countries from the Oapen record.")
countries_oapen = []
with open(oapen_download_file, 'r') as infile:
  # read the file as a dictionary for each row ({header : value})
  reader = csv.reader(infile)
  data = []
  for row in reader:
      data.append(row[0])
  countries_oapen = data[1:]
  # remove duplicates
  countries_oapen = list(set(countries_oapen))

# open jstor Excel workbook and create country list
# Excel-rows 1 to 8 are extra information, and need to be discarded. Excel row 9 contains column headers.
print()
print("Making a list of countries from the jstor record.")
countries_jstor = []
jstor_workbook = load_workbook(os.path.join(path_input_output,'rd_jstor.xlsx'), data_only=True)
jstor_worksheet = jstor_workbook['data']
for row in jstor_worksheet.values:
   countries_jstor.append(row[0])
countries_jstor = countries_jstor[1:]
# remove duplicates
countries_jstor = list(set(countries_jstor))

# open muse Excel workbook and create country list
countries_muse = []
print()
print("Making a list of countries from the Project Muse records.")
muse_workbook = load_workbook(os.path.join(path_input_output,'rd_muse.xlsx'), data_only=True)
muse_worksheet = muse_workbook['data']
for row in muse_worksheet.values:
   countries_muse.append(row[1])
countries_muse = countries_muse[1:]
countries_muse = list(set(countries_muse))

countries_oapen_new = []
for country in countries_oapen:
    if country not in countries_oapen_reference:
        countries_oapen_new.append(country)

countries_jstor_new = []
for country in countries_jstor:
    if country not in countries_jstor_reference:
        countries_jstor_new.append(country)

countries_muse_new = []
for country in countries_muse:
    if country not in countries_muse_reference:
        countries_muse_new.append(country)

# writing the unmatched/new countries to a .txt file
if len(countries_oapen_new) > 0 or len(countries_jstor_new) > 0 or len(countries_muse_new) > 0:
    new_countries_file_txt = os.path.join(path_input_output, "countries_new.txt")
    print()
    print("New countries found. Writing to a .txt file.")
    print(f"File: \n{new_countries_file_txt}")
    print()
    print("Please add these these new countries to the 'countries_combined.csv' file.")
    print("Make sure to add each to the right column (i.e. for Oapen, Jstor, or Project Muse)")
    print("Match them to the correct countries in other columns, and especially the country codes.")
    print("If the right country code is not yet in the document, look it up online.")
    print("Look for the two-letter ISO Alpha-2 codes.")
    print("You can find these (for example) on https://www.iban.com/country-codes.")
    print()
    # check which lists contain more than 0 items
    # write those with anything in them to the file
    with open (new_countries_file_txt, 'w') as file: 
        if len(countries_oapen_new) > 0:
            file.write("\nNew countries from the Oapen records: \n")
            for item in countries_oapen_new:
                file.write(f"{item}\n")
        if len(countries_jstor_new) > 0:
            file.write("\nNew countries from the Jstor records: \n")
            for item in countries_jstor_new:
                file.write(f"{item}\n")
        if len(countries_muse_new) > 0:
            file.write("\nNew countries from the Project Muse records: \n")
            for item in countries_muse_new:
                file.write(f"{item}\n")
else: 
    print()
    print("No new countries found in records. The 'countries_combined.csv' file does not need to be updated.")

print()
input("Script finished. Press Enter to close.\n")