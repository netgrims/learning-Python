from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
import time, os, csv, re
import pandas as pd

def download_wait(directory, timeout, nfiles=None):
    """
    Wait for downloads to finish with a specified timeout.

    Args
    ----
    directory : str
        The path to the folder where the files will be downloaded.
    timeout : int
        How many seconds to wait until timing out.
    nfiles : int, defaults to None
        If provided, also wait for the expected number of files.

    """
    seconds = 0
    dl_wait = True
    while dl_wait and seconds < timeout:
        time.sleep(0.5)
        dl_wait = False
        files = os.listdir(directory)

        if nfiles != None:
            if nfiles and len(files) != nfiles:
                dl_wait = True

        for fname in files:
            if fname.endswith('.crdownload'):
                dl_wait = True

        seconds += 0.5
    return seconds

def clean_filename(filename):
    # str(file).replace(":","-")
    # clean_string = re.sub("[^0-9a-zA-Z\s]+", "", filename)
    # clean_string.lower()
    clean_string = ''.join(letter for letter in filename if letter.isalnum())
    return clean_string

def rename_latest_download(download_file_path):
    os.chdir(download_file_path)
    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
    latest_file = os.path.join(download_file_path, files[-1])
    newname = clean_filename(book_title) + "_" + str(match_no) + ".csv"
    os.rename(latest_file, newname)
    print(f"{latest_file} (file no. {len(files)}) renamed to {newname}")

def check_exists_by_xpath(xpath):
    try:
        driver.find_element(By.XPATH, xpath)
    except NoSuchElementException:
        return False
    return True

# path settings
current_year = date.today().year
last_year = current_year - 1
homedir = os.path.expanduser('~')
path_script = os.path.join(os.path.dirname(__file__))
path_year = os.path.join(os.path.dirname(__file__), "oapen")
path_input_output = os.path.join(path_script, 'input-output')

print("To use this script, make sure you have the following files in the input-output folder of the script directory:")
print(" - The book_titles_merged.csv file.")
print()
print("Also, make sure you have Google Chrome installed, and that the Chrome WebDriver file is located in the script directory:")
print()
print(f"The script directory is {path_script}")
print(f"The input-output folder is {path_input_output}")
print()
input("If you are sure that these requirements are met, press Enter to continue. \n")
print()
print("Starting script.")

# checking if the target folder exists, creating it otherwise
if not os.path.exists(path_year):
    os.makedirs(path_year)

# selenium driver settings
chrome_options = Options()
chrome_options.add_argument("--disable-search-engine-choice-screen")
prefs = {"download.default_directory":path_year}
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

# open Excel workbook and create book list
# original_wb = load_workbook(os.path.join(path_script,'readership_data.xlsx'), data_only=True)
# original_ws = original_wb['readership_data']
# book_list = original_ws['A1':'A2']

# create new Excel workbook for writing and add headers
# new_wb = Workbook()
# new_ws = new_wb.active
# row = ['book_title', 'data_url', str(last_year) + '-total', str(last_year) + '-01', str(last_year) + '-02', str(last_year) + '-03', str(last_year) + '-04', str(last_year) + '-05', str(last_year) + '-06', str(last_year) + '07', str(last_year) + '08', str(last_year) + '09', str(last_year) + '10', str(last_year) + '11', str(last_year) + '12']
# new_ws.append(row)

# load book titles csv file (generated with the previous python script)
print()
print("Loading the book titles list from the file.")
merged_csv_path = os.path.join(path_input_output, "book_titles_merged.csv")
try: 
    # open the file in read mode
    merged_csv_file = open(merged_csv_path, 'r', encoding='utf-16')
except: 
    print()
    print("Problem loading book_titles_merged.csv file.")
    print(f"""
        Please check the following:
          1) The book_titles_merged.csv is in the correct path: {path_year}
          2) You cannot have the book_titles_merged.csv opened in another program (e.g. Excel or LibreOffice). If you do, please close it.
          3) The book_titles_merged.csv file is not empty.
          4) The book_titles_merged.csv file has the right name.
          """)

# open the file in read mode
merged_csv_file = open(merged_csv_path, 'r', encoding='utf-16')
# creating dictreader object
file = csv.DictReader(merged_csv_file)

# create an empty list
book_titles_oap_filtered = []
 
# iterate over each row and append values to empty list
for col in file:
    book_titles_oap_filtered.append(col['oapen title'])

print()
print("Book titles list loaded.")

# specify the url to use (oapen) for selenium
url = 'https://dashboard.oapen.org/#/dashboard'

# initialise selenium driver
print()
print("Initialising Selenium WebDriver.")
driver.maximize_window()
driver.get(url)

# login to oapen
print()
print("Logging in to oapen.")
username = driver.find_element(By.XPATH, '//*[@id="username"]')
username.clear()
# username redacted
username.send_keys("")
password = driver.find_element(By.XPATH, '//*[@id="password"]')
password.clear()
# password redacted
password.send_keys("")
login_button = driver.find_element(By.XPATH, '/html/body/div/table/tbody/tr/td/form/div/table/tbody/tr[3]/td/button')
login_button.click()

# scroll back to the top of the oapen webpage
driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.CONTROL + Keys.HOME)

# select per month and country view
print()
print("Selecting 'per month and country' view.")
wait = WebDriverWait(driver, 10)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/main/div/div/div/header/div/button/span/i'))).click()
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div[2]/main/div/div/div/aside/div[1]/div/div[3]/div[2]/div'))).click()

# select December last year as the End Month
print()
print("Selecting the right end month.")
last_year_dec = str(last_year) + "-12"
xpath_last_year_dec = "//div[@class='v-list-item v-list-item--link theme--light' and contains(.,'"+ last_year_dec + "')]"
# "//*[contains(text(), '" + last_year_dec + "')]"
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/main/div/div/div/div/div[1]/div/div[1]/div/div/div[1]/div[1]/div[1]/div'))).click()
wait.until(EC.element_to_be_clickable((By.XPATH, xpath_last_year_dec))).click()

# select Book as the Item Type
print()
print("Selecting the book item type.")
wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='v-select__slot' and contains(.,'Item type')]"))).click()
wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='v-list-item__title' and contains(.,'Book')]"))).click()

# make an empty list to add duplicate matches to 
# (this gets written to a file so that it can be reported to oapen)
duplicate_matches = []

# make an empty list to add downloaded titles to
downloaded_titles = []

# Search for the book title
print()
print("Searching for the books from the list.")
for book_title in book_titles_oap_filtered:
    # search for book
    print()
    print("Searching for " + book_title)
    driver.find_element(By.XPATH,'//*[contains(@id, "input-") and (@autocomplete="off") and not((@readonly="readonly"))]').click()
    driver.find_element(By.XPATH,'//*[contains(@id, "input-") and contains (@placeholder, "Start typing")]').clear()
    driver.find_element(By.XPATH,'//*[contains(@id, "input-") and contains (@placeholder, "Start typing")]').send_keys(book_title)

    # make a list of elements matching the book title in the search bar (for duplicates)
    match_xpath = "//span[text()='" + book_title + "']"
    # book_matches = wait.until(EC.presence_of_all_elements_located(((By.XPATH, "//*[contains(text(),'"+ book_title + "')]/../../.."))))
    book_matches = wait.until(EC.presence_of_all_elements_located(((By.XPATH, "//*[contains(text(),'"+ book_title + "')]/../../.."))))
    match_no = 0
    id = book_matches[0].get_attribute("id")

    # book_matches = book_matches[:2]

    # go through the list of matches
    for match in book_matches:
        wait.until(EC.element_to_be_clickable((By.XPATH,'//*[contains(@id, "input-") and (@autocomplete="off") and not((@readonly="readonly"))]'))).click()
        match_id = match.get_attribute("id")
        # wait.until(EC.element_to_be_clickable((By.ID, match_id))).click()
        # check if the match really is exactly the same book, and not another book with a similar name
        # oapen_page_title = driver.find_element(By.XPATH, "//*[contains(text(),'"+ book_title + "') and not (@class='v-list-item__title')]")
        # oapen_page_title = driver.find_element(By.XPATH, "//*[text()='"+ book_title + "' and not (@class='v-list-item__title')]")
        # if oapen_page_title.text == book_title:
        wait.until(EC.element_to_be_clickable(match)).click()

        # check if there is data (.csv) file available for this entry, and download if applicable
        time.sleep(1)
        if check_exists_by_xpath("//*[contains(text(), 'No data available')]") == False:
            print(f"{book_title}: .csv file available.")

            # check if the match is an exact match with the book title, download if so
            title_match = False
            try:
                oapen_page_title = driver.find_element(By.XPATH, "//*[text()='"+ book_title + "' and not (@class='v-list-item__title')]")
                print(f"Checking if {oapen_page_title.text} matches with {book_title}")
                if oapen_page_title.text == book_title:
                    title_match = True
            except NoSuchElementException:
                print("No match found.")
            
            if title_match == True:
                print(f"{oapen_page_title.text} matches {book_title}. Downloading file.")
                # download csv file. FYI: some of the duplicates have no associated data, so clicking download also doesn't do anything.
                # driver.find_element(By.XPATH,'//*[contains(@id,"json-to-csv-")]/button').click()
                download_button = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[contains(@id,"json-to-csv-")]/button')))
                download_button.click()

                # wait for download complete
                match_no += 1
                print(f"Download for {book_title} (match {match_no}) completed in {str(download_wait(path_year,3))} seconds.")

                # add to list of downloaded books
                downloaded_titles.append(book_title)

                # rename the file
                rename_latest_download(path_year)

                # check if match is a duplicate
                # if match is a duplicate, merge it with the last file
                if match_no == 2:
                    print()
                    print("Duplicate found. Merging: " + book_title)

                    # add the titles of duplicate items on oapen to a list for future reference and to report to oapen
                    duplicate_matches.append(book_title)

                    # merge duplicate files
                    file_01 = os.path.join(clean_filename(book_title) + "_" + str(match_no - 1) + ".csv")
                    file_02 = os.path.join(clean_filename(book_title) + "_" + str(match_no) + ".csv")
                    df1 = pd.read_csv(file_01)
                    print(df1.head())
                    df2 = pd.read_csv(file_02)
                    print(df2.head())
                    headers = list(df1.columns)
                    print(headers)
                    # set the common columns as the index
                    # df1 = df1.set_index(headers[0:5])
                    # df2 = df2.set_index(headers[0:5])
                    # df1 = df1.set_index(headers)
                    # df2 = df2.set_index(headers)
                    # merged_df = pd.merge(df1, df2, on='country')
                    # merged_df = df1 + df2
                    # merged_df = df1.merge(df2, on="outer")
                    merged_df = pd.concat([df1, df2]).groupby(['country', 'countryCode', 'latitude', 'longitude', 'yearMonth']).sum().reset_index()
                    print(merged_df.head())
                    #file_01.close()
                    #file_02.close()

                    # remove old duplicate file
                    os.remove(file_02)
                    os.remove(file_01)

                    # write new merged file
                    merged_df.to_csv(file_01, index=False)
                    #with open(file_01, 'w', newline='', encoding='utf-8') as file:
                    #    writer = csv.writer(file)
                    #    for row in merged_df:
                    #        writer.writerow(row)
                    #    file.close()
                    print("Duplicates merged: " + file_01)

# closing selenium webdriver
driver.quit()

print()
print(f"Downloads finished. Look in {path_year} for the .csv files per book.")

# write duplicate matches to .txt file
duplicate_matches_file_path = os.path.join(path_input_output, "oapen_duplicate_matches.txt")
with open (duplicate_matches_file_path, 'w') as file: 
    for item in duplicate_matches:
        file.write(f"{item}\n")

# print message reporting on duplicate matches
print()
if len(duplicate_matches) > 0:
    print("Duplicate matches for book titles in Oapen found.")
    print(f"A list of duplicate titles has been saved in {duplicate_matches_file_path}")
    print("You can use this list to report duplicates to Oapen.")

# make list of non-downloaded titles
non_downloaded_titles = []
for title in book_titles_oap_filtered:
    if title not in downloaded_titles:
        non_downloaded_titles.append(title)

# write list of non-downloaded titles to file
print()
print("Checking for non-downloaded files.")
non_downloaded_titles_file_path = os.path.join(path_input_output, "oapen_non-downloaded_titles.txt")
with open (non_downloaded_titles_file_path, 'w') as file: 
    for item in non_downloaded_titles:
        file.write(f"{item}\n")

# print message reporting on non-downloaded titles
print()
if len(non_downloaded_titles) > 0:
    print(f"""
Some titles were not downloaded.
    A list of non-downloaded titles has been saved to {non_downloaded_titles_file_path}.
    Please check this list to make sure no error was made, and download files manually when necessary.
    Rename any manually downloaded files by removing all spaces and special characters. This is necessary for the next steps.
          """)

print()
input("Script finished. Press enter to close.")