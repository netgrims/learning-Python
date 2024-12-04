import csv, time, re, os
import pandas as pd
import jellyfish as jf
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys

# path settings
current_year = date.today().year
last_year = current_year - 1
homedir = os.path.expanduser('~')
path_script = os.path.join(os.path.dirname(__file__))
path_year = os.path.join(os.path.dirname(__file__), str(last_year) + "_oapen")
path_input_output = os.path.join(path_script, 'input-output')

print("To use this script, make sure you have the following files in the input-output folder of the script directory:")
print(" - The JStor readership data excel file, renamed to rd_jstor.xlsx")
print(" - The Project Muse readership data excel file, renamed to rd_muse.xlsx")
print()
print(f"The path to the input-output folder is: \n{path_input_output}")
print()
input("If you are sure that these requirements are met, press Enter to continue. \n")
print()
print("Starting script.")

# check if there is a combined countries file present
countries_combined_file = os.path.join(path_input_output, "countries_combined.csv")
try:
    with open(countries_combined_file, 'r', newline='', encoding='utf-16') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_STRINGS)
        existing_merged_file = csv.DictReader(file)
        countries_oapen_reference = []
        countries_jstor_reference = []
        countries_muse_reference = []
        # write the contents for each column to lists per database
        for row in existing_merged_file:
            countries_oapen_reference.append(row['countries_oapen'])
            countries_jstor_reference.append(row['countries_jstor'])
            countries_muse_reference.append(row['countries_muse'])
        file.close()
except FileNotFoundError:
    print()
    print("No existing countries_combined.csv file found in the directory.") 
    print("If you want to use an existing list, please make sure the file is there, and restart the script.")
    input("Otherwise, press enter to continue. This will create a new countries_combined.csv file. \n")
    with open(countries_combined_file, 'w', newline='', encoding='utf-16') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_STRINGS)
        writer.writerow(["country_code", "countries_oapen", "countries_jstor", "countries_muse"])
        file.close()

# remove empty strings from lists
while '' in countries_oapen_reference:
    countries_oapen_reference.remove('')
while '' in countries_jstor_reference:
    countries_jstor_reference.remove('')
while '' in countries_muse_reference:
    countries_muse_reference.remove('')

# selenium driver settings
chrome_options = Options()
chrome_options.add_argument("--disable-search-engine-choice-screen")
prefs = {"download.default_directory":path_input_output}
chrome_options.add_experimental_option("prefs", prefs)
driver = webdriver.Chrome(options=chrome_options)
wait = WebDriverWait(driver, 10)

# specify the url to use (oapen)
url = 'https://dashboard.oapen.org/#/dashboard'

# initialise selenium driver
print()
print("Opening oapen website.")
driver.maximize_window()
driver.get(url)

# login to oapen
# if the username and/org password ever change, you can edit them here
print()
print("Logging in to oapen.")
username = driver.find_element(By.XPATH, '//*[@id="username"]')
username.clear()
# username redacted
username.send_keys("")
password = driver.find_element(By.XPATH, '//*[@id="password"]')
password.clear()
# password redacted
password.send_keys("")
login_button = driver.find_element(By.XPATH, '/html/body/div/table/tbody/tr/td/form/div/table/tbody/tr[3]/td/button')
login_button.click()

# select by month view
print()
print("Selecting 'per month and country' view.")
wait = WebDriverWait(driver, 10)
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div/main/div/div/div/header/div/button/span/i'))).click()
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div[2]/main/div/div/div/aside/div[1]/div/div[3]/div[2]/div'))).click()

# show all records on one page
print()
print("Showing all records on one page.")
wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="main"]/div[1]/main/div/div/div/div/div[2]/div/div/div/div[3]/div/div[2]/div[1]/div/div/div'))).click()
driver.find_element(By.XPATH, '//*[contains(@id, "-4") and (@tabindex="0") and (.//text()= "All")]').click()

# scroll back to the top of the oapen webpage
driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.CONTROL + Keys.HOME)
time.sleep(0.1)

# download the .csv file
print()
print("Downloading .csv file.")
download_button = wait.until(EC.element_to_be_clickable((By.XPATH,'//*[contains(@id,"json-to-csv-")]/button')))
download_button.click()

time.sleep(4)

# close the selenium webdriver
print()
print("Closing Selenium Webdriver")
driver.quit()

def clean_filename(filename):
    # str(file).replace(":","-")
    clean_string = re.sub("[^0-9a-zA-Z\s]+", "", filename)
    clean_string.lower()
    return clean_string

def rename_latest_download(download_file_path):
    os.chdir(download_file_path)
    files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
    latest_file = os.path.join(download_file_path, files[-1])
    newname = clean_filename(book_title) + "_" + str(match_no) + ".csv"
    os.rename(latest_file, newname)
    print(latest_file + + " (file no. " + len(files) + ") renamed to " + newname)

# open oapen csv and create country list
print()
print("Making a list of countries from the Oapen record.")
oapen_download_file = "monthly_requests_per_country.csv"
oapen_download_file = os.path.join(path_input_output, oapen_download_file)
countries_oapen = {}
try: 
    # open the file in read mode
    merged_csv_file = open(oapen_download_file, 'r')
    oapen_file = csv.reader(merged_csv_file)
    oapen_file = oapen_file[1:]
    # make a list from the first column (labeled 'country')
    for item in oapen_file:
        countries_oapen.append(item['country'])
except: 
    print("Problem loading monthly_requests_per_country.csv file.")
    print(f"""
        Please check the following:
          1) The monthly_requests_per_country.csv is in the correct path: {path_input_output}
          2) You cannot have the monthly_requests_per_country.csv opened in another program (e.g. Excel or LibreOffice). If you do, please close it.
          3) The monthly_requests_per_country.csv file is not empty.
          4) The monthly_requests_per_country.csv file has the right name.

        Close this script, and run it again from the start.
          """)

countries_oapen = []

# open the oapen file in universal line ending mode 
with open(oapen_download_file, 'r') as infile:
  # read the file as a dictionary for each row ({header : value})
  #reader = csv.DictReader(infile)
  reader = csv.reader(infile)
  data = []
  for row in reader:
    data.append(row[0])
    # for header, value in row.items():
      #try:
         #data[header].append(value)
      #except KeyError:
         #data[header] = [value]
  countries_oapen = data[1:]

print()
print("COUNTRIES OAPEN:")
print(countries_oapen)

# input("Press Enter to continue. \n")

countries_jstor = []

# open jstor Excel workbook and create country list
# Excel-rows 1 to 8 are extra information, and need to be discarded. Excel row 9 contains column headers.
print()
print("Making a list of countries from the jstor record.")
jstor_workbook = load_workbook(os.path.join(path_input_output,'rd_jstor.xlsx'), data_only=True)
jstor_worksheet = jstor_workbook['Sheet1']
for row in jstor_worksheet.values:
   countries_jstor.append(row[0])
# countries_jstor = jstor_worksheet['A']
countries_jstor = countries_jstor[1:]
countries_jstor = list(set(countries_jstor))

print()
print("COUNTRIES JSTOR:")
print(countries_jstor)

# input("Press Enter to continue. \n")

# open muse Excel workbook and create country list
countries_muse = []
print()
print("Making a list of countries from the Project Muse records.")
muse_workbook = load_workbook(os.path.join(path_input_output,'rd_muse.xlsx'), data_only=True)
muse_worksheet = muse_workbook['pivot']
for row in muse_worksheet.values:
   countries_muse.append(row[0])
countries_muse = countries_muse[1:]
countries_muse = list(set(countries_muse))

print()
print("COUNTRIES MUSE:")
print(countries_muse)

# input("Press Enter to continue. \n")

# define process for fuzzy-matching two lists
def get_closest_match(x, list_random, threshold=0):
    best_match = None
    highest_jaro_wink = 0
    for current_string in list_random:
        current_score = jf.jaro_winkler_similarity(x, current_string)
        if(current_score > highest_jaro_wink):
            if current_score > threshold:
              highest_jaro_wink = current_score
              best_match = current_string
    return best_match

# make a list of lists of the exact matches in the three lists
print("Gathering exact duplicates into a list of lists.")
countries_combined = []
countries_jstor_unique = countries_jstor
countries_muse_unique = countries_muse
for country_oapen in countries_oapen:
  if country_oapen in countries_jstor:
    countries_jstor_unique.remove(country_oapen)
    if country_oapen in countries_muse:
        countries_combined.append([country_oapen, country_oapen, country_oapen])
        countries_muse_unique.remove(country_oapen)
    else: 
        countries_combined.append([country_oapen, country_oapen, ''])
  else:
    if country_oapen in countries_muse:
        countries_combined.append([country_oapen, '', country_oapen])
        countries_muse_unique.remove(country_oapen)
    else: 
        countries_combined.append([country_oapen, '', ''])

countries_jstor_unique_unmatched = countries_jstor_unique
countries_muse_unique_unmatched = countries_muse_unique

print()
print("Fuzzy-matching unique countries.")
for country in countries_combined:
  if country[0] != '':
    if country[1] == '':
      country[1] = get_closest_match(country[0], countries_jstor_unique, 0.4)
      if country[2] != None:
        countries_jstor_unique_unmatched.remove(country[1])
    if country[2] == '':
      country[2] = get_closest_match(country[0], countries_muse_unique, 0.4)
      if country[2] != None:
        countries_muse_unique_unmatched.remove(country[2])

# report on unmatched countries and append them to the combined list
print()
print(f"{len(countries_jstor_unique_unmatched)} countries from the JStor list could not be matched.")
for item in countries_jstor_unique_unmatched:
   countries_combined.append(['', item, ''])

print(f"{len(countries_muse_unique_unmatched)} countries from the Project Muse list could not be matched.")
for item in countries_muse_unique_unmatched:
   countries_combined.append(['','',item])

# Writing the output to a .csv file.
print()
print("Writing the output to a .csv file.")
countries_combined_file = os.path.join(path_input_output, "countries_combined.csv")
with open(countries_combined_file, 'w', newline='', encoding='utf-16') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_STRINGS)
    header = ['countries_oapen', 'countries_jstor', 'countries_muse']
    writer.writerow(header)
    writer.writerows(countries_combined)
    f.close()

# Writing the unmatched strings to files: jstor
print()
print("Writing the unmatched jstor countries to a .txt file.")
countries_jstor_unmatched_file_txt = os.path.join(path_input_output, "countries_jstor_unmatched.txt")
with open (countries_jstor_unmatched_file_txt, 'w') as file: 
    for item in countries_jstor_unique_unmatched:
        file.write(f"{item}\n")

# Writing the unmatched strings to files: project muse
print()
print("Writing the unmatched project muse countries to a .txt file.")
countries_muse_unmatched_file_txt = os.path.join(path_input_output, "countries_muse_unmatched.txt")
with open (countries_muse_unmatched_file_txt, 'w') as file: 
    for item in countries_muse_unique_unmatched:
        file.write(f"{item}\n")

print()
input("Script finished. Press Enter to close.")